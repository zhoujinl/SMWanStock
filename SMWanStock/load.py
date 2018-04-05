# -*- coding:utf-8 -*-

from openpyxl import load_workbook
from openpyxl import Workbook

class Load:

    """ 从excel读取数据"""

    def __init__(self, filename):
        self.wb = load_workbook(filename)
        self.ws =  self.wb["Sheet01"]
        self.rowList = []

    def read(self):

        for row in self.ws.iter_rows(min_row=2):  # 去掉标题
            rs = ''
            for cell in row:
                rs  = rs +  cell.value  +  ','
            rs = str(rs[:-1])
            self.rowList.append(rs)

    def close(self):
        self.wb.close()

if __name__ == "__main__":
    ld = Load(r"E:\Git\SMWanStock\tests\data\stock.xlsx")
    ld.read()
    alt = ld.rowList
    print(len(alt))
    lt = ['B,600691,107000,3.95,Tue Jan  2 15:23:39 2018', 'B,000807,13900,10.62,Tue Jan  2 15:23:39 2018']

    for i in range(len(lt),len(alt)):

          print(alt[i])

    print(cmp(alt,lt))